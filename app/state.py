import calendar
import decimal
import re
from datetime import datetime, timedelta, date, time

from dateutil.relativedelta import relativedelta
from django.db.models import Sum, ExpressionWrapper, F, FloatField, Count, IntegerField, Q
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, LineChart
from openpyxl.chart.axis import NumericAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from .models import *

wb = Workbook()


def transform_data_heart_map(data_dict: dict) -> list:
    """Преобразует данные в формат, подходящий для тепловой карты."""

    # Функция для извлечения числового значения из ключа месяца
    def extract_month_number(month_key):
        match = re.match(r'(\d+)', str(month_key))
        if match:
            return int(match.group(1))
        return float('inf')  # Если не удалось, ставим бесконечность для сортировки в конец

    # Функция для извлечения числового значения из ключа покупки
    def extract_purchase_number(purchase_key):
        match = re.match(r'(\d+)', purchase_key)
        if match:
            return int(match.group(1))
        return -1  # Если не удалось, ставим -1 для сортировки в начало

    # 1. Извлекаем и сортируем месяцы по возрастанию их числового значения
    months = sorted(data_dict.keys(), key=extract_month_number)

    # 2. Извлекаем все уникальные типы покупок
    purchase_set = set()
    for month_data in data_dict.values():
        purchase_set.update(month_data.keys())

    # 3. Сортируем покупки по убыванию их числового значения
    purchases = sorted(purchase_set, key=extract_purchase_number, reverse=True)

    # 4. Формируем строки с данными
    rows = []
    for purchase in purchases:
        row = [purchase]
        for month in months:
            value_data = data_dict.get(month, {}).get(purchase, {'count': 0, 'users': []})
            count = value_data['count']
            users = value_data['users']
            row.append(
                {'count': count, 'users': users})  # Сохраняем данные о пользователях для добавления комментария позже
        rows.append(row)

    # 5. Формируем заголовок
    header = ['Кол-во покупок/Месяцев с последней покупки'] + months

    # 6. Добавляем заголовок в конец списка
    rows.append(header)

    return rows


def heart_map_data():
    """Собирает данные о покупках пользователей и формирует словарь для тепловой карты."""
    users = Client.objects.all()
    data_dict = {}
    for user in users:
        if user.chat_id == '980964972':
            continue  # Пропускаем указанного пользователя

        orders = user.orders_for_user.filter(type__in=['buy', 'not_find_product'], pay_status='complite')
        orders_count = orders.count()
        orders_count_str = f'{orders_count} Покупка' if orders_count == 1 else f'{orders_count} Покупки'

        if orders_count > 0:
            date_obj = orders.first().time
            try:
                date_obj = datetime(date_obj.year, date_obj.month, date_obj.day)
                naive_datetime = datetime.combine(date_obj, time.min)
                aware_datetime = timezone.make_aware(naive_datetime)
                n = timezone.now() - aware_datetime
            except Exception as e:
                print(f"Ошибка при обработке времени для пользователя {user.chat_id}: {e}")
                continue  # Пропускаем пользователя, если возникла ошибка

            month = 0
            try:
                month += n.days // 30
            except Exception as e:
                pass
            try:
                month += n.month
            except Exception as e:
                pass

            month_str = f'{month} Месяц' if month == 1 else f'{month} Месяцев'

            if month_str in data_dict:
                if orders_count_str in data_dict[month_str]:
                    data_dict[month_str][orders_count_str]['count'] += 1
                    data_dict[month_str][orders_count_str]['users'].append(user.chat_id)
                else:
                    data_dict[month_str][orders_count_str] = {'count': 1, 'users': [user.chat_id]}
            else:
                data_dict[month_str] = {orders_count_str: {'count': 1, 'users': [user.chat_id]}}

    return data_dict


def heart_map(wb):
    """Создает тепловую карту в Excel."""
    ws = wb.active
    ws.title = 'Тепловая карта'

    # Данные для тепловой карты
    raw_data = heart_map_data()
    data = transform_data_heart_map(raw_data)

    # Запись данных в ячейки
    for row_index, row_data in enumerate(data, start=1):
        for col_index, cell_data in enumerate(row_data, start=1):
            if isinstance(cell_data, dict):
                # Если ячейка содержит словарь, это данные о количестве и пользователях
                count = cell_data['count']
                users = cell_data['users']
                ws.cell(row=row_index, column=col_index, value=count)  # Записываем только число
                if users:
                    # Добавляем комментарий с ID пользователей
                    comment_text = f"ID пользователей: {', '.join(users)}"
                    comment = Comment(comment_text, "System")  # Замените "System" на имя автора комментария
                    ws.cell(row=row_index, column=col_index).comment = comment
            else:
                # Если это не словарь, это заголовок или название покупки
                ws.cell(row=row_index, column=col_index, value=cell_data)

    # Создаем правило условного форматирования ColorScale
    color_scale_rule = ColorScaleRule(start_type='min',
                                      start_color='00CC00',
                                      mid_type='percentile',
                                      mid_value=25,
                                      mid_color='FFFF33',
                                      end_type='max',
                                      end_color='FFFF0000')

    # Применяем правило ко всему диапазону данных.
    # Важно! Измените диапазон, чтобы он соответствовал вашим данным.
    # Здесь предполагается, что данные начинаются с B1 и идут до столбца, соответствующего кол-ву месяцев.
    data_range = f'B1:MM1000'  # 65 = 'A', chr(65 + num_months) - буква столбца конца диапазона
    ws.conditional_formatting.add(data_range, color_scale_rule)


def revenue_structure_data(startdate, enddate):
    orders = Order.objects.filter(close_time__range=(startdate, enddate), type__in=['buy', 'not_find_product'],
                                  pay_status='complite')
    n = {}
    total_buy = 0
    total_income = 0
    for order in orders:
        inc = order.total_product_price - order.product_price
        total_buy += 1
        total_income += inc
        if order.name in n:
            n[order.name]['buy'] += 1
            n[order.name]['income'] += inc
        else:
            n[order.name] = {'buy': 1, 'income': inc}

    data = []
    for i in n:
        lst = [i, round(n[i]['income'] / total_income * 100, 2), round(n[i]['buy'] / total_buy * 100, 2),
               n[i]['income']]
        data.append(lst)
    return data


def revenue_structure(wb, startdate, enddate):
    ws = wb.create_sheet("Структура выручки")

    data = [
               ["", "% от выручки", 'Количество сделок в процентах', 'Выручка в $'],
           ] + revenue_structure_data(startdate, enddate)

    for row in data:
        ws.append(row)

    pie1 = PieChart()
    pie2 = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(data))
    data_values1 = Reference(ws, min_col=2, min_row=2, max_row=len(data))
    data_values2 = Reference(ws, min_col=3, min_row=2, max_row=len(data))

    pie1.add_data(data_values1)
    pie2.add_data(data_values2)

    pie1.set_categories(labels)
    pie2.set_categories(labels)

    pie1.title = "% От выручки"

    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showPercent = True  # Показывать проценты
    pie1.dataLabels.showVal = False  # Отключить абсолютные значения
    pie1.dataLabels.showCatName = False  # Отключить подписи категорий
    pie1.dataLabels.showSerName = False  # Отключить подпись "Ряд 1"

    pie2.title = "% От продаж"
    pie2.dataLabels = DataLabelList()
    pie2.dataLabels.showPercent = True
    pie2.dataLabels.showVal = False
    pie2.dataLabels.showCatName = False
    pie2.dataLabels.showSerName = False

    ws.add_chart(pie1, "F1")
    ws.add_chart(pie2, "P1")


def iterate_days(start_date, end_date):
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += relativedelta(months=1)


def statistics_for_each_service_data(startdate, enddate):
    dates = [i for i in iterate_days(startdate, enddate)]
    data1 = [
        [None, ] + dates,
    ]
    data2 = [
        [None] + dates,
    ]

    data3 = [
        [None],
        [None],
        [None],
        [None],
        [None],
        ['Название сервиса', 'кол-во продаж', 'Выручка']
    ]
    n = {}
    for product in Products.objects.all():
        prod_name = product.name
        prod1 = [prod_name]
        prod2 = [prod_name]
        n[prod_name] = {
            'count': 0,
            'inc': 0
        }
        for dt in dates:
            first_day = date(dt.year, dt.month, 1)
            last_day = date(dt.year, dt.month, calendar.monthrange(dt.year, dt.month)[1])
            orders = Order.objects.filter(name=prod_name, close_time__range=(first_day, last_day),
                                          type__in=['buy', 'not_find_product'],
                                          pay_status='complite')
            if orders:
                total_purchases = orders.count()
                total_revenue = orders.aggregate(
                    total_revenue=Sum(F('total_product_price') - F('product_price'))
                )['total_revenue']
                n[prod_name]['count'] += total_purchases
                n[prod_name]['inc'] += total_revenue
                prod1.append(total_purchases)
                prod2.append(total_revenue or 0)
            else:
                prod1.append(0)
                prod2.append(0)
        data1.append(prod1)
        data2.append(prod2)
    for i in n:
        data3.append([i, n[i]['count'], n[i]['inc']])
    return data1, data2, data3


def statistics_for_each_service(wb, startdate, enddate):
    ws = wb.create_sheet("Статистика для каждого сервиса")
    data1, data2, data3 = statistics_for_each_service_data(startdate, enddate)

    # Запишем данные построчно на лист
    for row in data1:
        ws.append(row)

    for row in data2:
        ws.append(row)

    for row in data3:
        ws.append(row)

    # Создаём линейный график
    chart1 = LineChart()
    chart1.title = "Кол-во сделок"

    chart2 = LineChart()
    chart2.title = "Прибыль"

    cats1 = Reference(ws, min_col=1, max_col=len(data1[0]), min_row=1, max_row=1)
    cats2 = Reference(ws, min_col=1, max_col=len(data1[0]), min_row=len(data1) + 1, max_row=len(data1) + 1)
    chart1.set_categories(cats1)
    chart2.set_categories(cats2)

    data_ref1 = Reference(ws, min_col=1, max_col=len(data1[0]), min_row=2, max_row=len(data1))
    data_ref2 = Reference(ws, min_col=1, max_col=len(data1[0]), min_row=2 + len(data1), max_row=len(data1) * 2)
    chart1.add_data(data_ref1, from_rows=True, titles_from_data=True)
    chart2.add_data(data_ref2, from_rows=True, titles_from_data=True)

    chart1.y_axis = NumericAxis()
    chart1.y_axis.majorTickMark = "cross"
    chart1.y_axis.delete = False

    chart2.y_axis = NumericAxis()
    chart2.y_axis.majorTickMark = "cross"
    chart2.y_axis.delete = False

    ws.add_chart(chart1, "I1")
    ws.add_chart(chart2, "I16")


def deal_stats_date(startdate, enddate):
    dates = [i for i in iterate_days(startdate, enddate)]
    data_1_new = [[], [], []]
    data_2_new = [[], [], []]
    data3 = [
        [None],
        [None],
        [None],
        [None],
        [None],
        ['Общее количество сделок', 'Количество сделок с новыми клиентами', 'Количество сделок со старыми клиентами',
         'Общая выручка', 'Выручка от новых клиетов', 'Выручка от старых клиентов'],
        [0, 0, 0, 0, 0, 0]
    ]
    for dt in dates:
        first_day = date(dt.year, dt.month, 1)
        last_day = date(dt.year, dt.month, calendar.monthrange(dt.year, dt.month)[1])
        orders = Order.objects.filter(close_time__range=(first_day, last_day), type__in=['buy', 'not_find_product'],
                                      pay_status='complite')
        total_purchases = orders.count()
        total_purchases_new_user = orders.filter(is_first_buy=True).count()
        total_purchases_old_user = orders.filter(is_first_buy=False).count()

        total_revenue = orders.aggregate(
            total_revenue=Sum(F('total_product_price') - F('product_price'))
        )['total_revenue']
        total_revenue_new_user = \
            orders.filter(is_first_buy=True).aggregate(
                Sum__total_product_price__sub=Sum(F('total_product_price') - F('product_price'))
            )[
                'Sum__total_product_price__sub']
        total_revenue_old_user = \
            orders.filter(is_first_buy=False).aggregate(
                Sum__total_product_price__sub=Sum(F('total_product_price') - F('product_price'))
            )[
                'Sum__total_product_price__sub']
        try:
            data3[-1][0] += total_purchases
            data3[-1][1] += total_purchases_new_user
            data3[-1][2] += total_purchases_old_user
            data3[-1][3] += total_revenue
            data3[-1][4] += total_revenue_new_user
            data3[-1][5] += total_revenue_old_user
        except Exception:
            pass

        data_1_new[0].append(total_purchases)
        data_1_new[1].append(total_purchases_new_user)
        data_1_new[2].append(total_purchases_old_user)

        data_2_new[0].append(total_revenue)
        data_2_new[1].append(total_revenue_new_user)
        data_2_new[2].append(total_revenue_old_user)

    return dates, data_1_new, data_2_new, data3


def deal_stats(wb, startdate, enddate):
    ws = wb.create_sheet("Статистика по сделкам")
    dates, data_1_new, data_2_new, data3 = deal_stats_date(startdate, enddate)
    data1 = [
        [None] + dates,
        ["Всего сделок"] + data_1_new[0],
        ["Сделки новых пользователей"] + data_1_new[1],
        ["Сделки старых пользователей"] + data_1_new[2],
    ]
    data2 = [
        [None] + dates,
        ["Общая выручка"] + data_2_new[0],
        ["Выручка от новых пользователей"] + data_2_new[1],
        ["Выручка от старых пользователей"] + data_2_new[2],
    ]

    # Запишем данные построчно на лист
    for row in data1:
        ws.append(row)

    for row in data2:
        ws.append(row)

    for row in data3:
        ws.append(row)

    # Создаём линейный график
    chart1 = LineChart()
    chart1.title = "Кол-во сделок"

    chart2 = LineChart()
    chart2.title = "Прибыль"

    cats1 = Reference(ws, min_col=1, max_col=len(data1[0]), min_row=1, max_row=1)
    cats2 = Reference(ws, min_col=1, max_col=len(data1[0]), min_row=5, max_row=5)

    chart1.set_categories(cats1)
    chart2.set_categories(cats2)

    data_ref1 = Reference(ws, min_col=1, max_col=len(data1[0]), min_row=2, max_row=4)
    data_ref2 = Reference(ws, min_col=1, max_col=len(data1[0]), min_row=6, max_row=8)
    chart1.add_data(data_ref1, from_rows=True, titles_from_data=True)
    chart2.add_data(data_ref2, from_rows=True, titles_from_data=True)

    chart1.y_axis = NumericAxis()
    chart1.y_axis.majorTickMark = "cross"
    chart1.y_axis.delete = False

    chart2.y_axis = NumericAxis()
    chart2.y_axis.majorTickMark = "cross"
    chart2.y_axis.delete = False

    ws.add_chart(chart1, "I1")
    ws.add_chart(chart2, "I16")


def mailing_stats_data(startdate, enddate):
    data = []
    for mail in Mailing.objects.filter(date__range=[startdate, enddate]):
        data.append([mail.date, mail.text, mail.send_users, mail.buy_users, mail.buy_summ])
    return data


def mailing_stats(wb, startdate, enddate):
    ws = wb.create_sheet("Статистика рассылки")
    data = [
               ["Дата рассылки", "Текст рассылки", "Кол-во получивших", "Купили за 3 дня", "Сумма покупок"]
           ] + mailing_stats_data(startdate, enddate)
    for row in data:
        ws.append(row)


def conversion_data(startdate, enddate):
    startdate = date(year=int(startdate.split('-')[0]), month=int(startdate.split('-')[1]),
                     day=int(startdate.split('-')[2]))
    enddate = date(year=int(enddate.split('-')[0]), month=int(enddate.split('-')[1]),
                   day=int(enddate.split('-')[2]))
    first = [0, 0, 0]
    second = [0, 0, 0]
    for user in Client.objects.filter(Q(join_time__range=[startdate, enddate]) |
                                      Q(stay_old__range=[startdate, enddate]) |
                                      Q(stay_new__range=[startdate, enddate])
                                      ):
        if user.stay_old and startdate <= user.stay_old <= enddate:
            first[0] += 1
            first[1] += 1
        elif user.stay_new and startdate <= user.stay_new <= enddate:
            first[0] += 1
            second[1] += 1
            second[0] += 1
        elif startdate <= user.join_time <= enddate:
            second[0] += 1
    try:
        first_persent = int(first[1] / first[0] * 100)
    except Exception:
        first_persent = 0
    try:
        second_persent = int(second[1] / second[0] * 100)
    except Exception:
        second_persent = 0
    first[2] = f'{first_persent}%'
    second[2] = f'{second_persent}%'

    return first, second


def conversion(wb, startdate, enddate):
    ws = wb.create_sheet("Конверсия")
    first, second = conversion_data(startdate, enddate)
    data = [
        ['Общее количество клиентов', 'Стали старыми', 'Процент'],
        first,
        [None, None, None],
        ['Кол-во вступивших в бота', 'Кол-во новых клиентов', 'Процент'],
        second
    ]
    for row in data:
        ws.append(row)


def active_users_data(startdate, enddate):
    dates = [i for i in iterate_days(startdate, enddate)]
    first = []
    second = []
    for active in Active_users.objects.filter(date__range=[startdate, enddate]):
        first.append(active.buy_users_count.all().count())
        second.append(active.user_action_count.all().count())
    return dates, first, second


def active_users(wb, startdate, enddate):
    ws = wb.create_sheet("Активные пользователи")
    dates, first, second = active_users_data(startdate, enddate)
    data = [
        [None] + dates,
        ["Совершили покупку"] + first,
        ["Нажали кнопку"] + second,
    ]
    for row in data:
        ws.append(row)
    chart = LineChart()
    chart.title = "Активность пользователей"

    cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=3)

    chart.set_categories(cats)

    data_ref = Reference(ws, min_col=2, max_col=len(data[0]), min_row=2, max_row=3)
    chart.add_data(data_ref, from_rows=True, titles_from_data=True)

    chart.y_axis = NumericAxis()
    chart.y_axis.majorTickMark = "cross"
    chart.y_axis.delete = False

    ws.add_chart(chart, "A5")


def balance_stat_data(startdate, enddate):
    dates = [i for i in iterate_days1(startdate, enddate)]
    balance_summ = []
    for balance in BalanceHistory.objects.filter(date__range=[startdate, enddate]):
        balance_summ.append(balance.total_balance)
    return dates, balance_summ


def iterate_days1(start_date, end_date):
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)


def balance_stat(wb, startdate, enddate):
    ws = wb.create_sheet("Статистика балансов")
    dates, balance_summ = balance_stat_data(startdate, enddate)
    data = [
        [None] + dates,
        ["Сумма балансов"] + balance_summ,
    ]
    for row in data:
        ws.append(row)

    chart = LineChart()
    chart.title = "Сумма балансов пользоватлей"

    data_ref = Reference(ws, min_col=1, max_col=len(data[0]), min_row=2, max_row=2)
    chart.add_data(data_ref, from_rows=True, titles_from_data=True)

    chart.y_axis = NumericAxis()
    chart.y_axis.majorTickMark = "cross"
    chart.y_axis.delete = False

    ws.add_chart(chart, "A5")


name = {
    'yt': 'Ютуб',
    'vk': 'Вконтакте',
    'tk': 'Тикток',
    'tv': 'Твитч',
    'tg': 'Телеграм',
    'ref': 'Реферальная система'
}


def traffic_state_data(type, startdate, enddate, all_user):
    clients = Client.objects.filter(join_time__range=[startdate, enddate], invite_ref__type=type)
    count = clients.count()
    first_buy = clients.filter(stay_new__range=[startdate, enddate]).count()
    try:
        conversion = int(first_buy / count * 100)
    except Exception:
        conversion = 0
    try:
        percent = int(count / all_user * 100)
    except Exception:
        percent = 0
    return [name[type], count, percent, first_buy, conversion]


def traffic_state(wb, startdate, enddate):
    clients = Client.objects.filter(join_time__range=[startdate, enddate],
                                    invite_ref__type__in=['yt', 'vk', 'tk', 'tv', 'tg', 'ref']).count()
    data = [
        [None, 'Количество вступлений', 'Процент вступлений', 'Стали новыми клиентами', 'Процент новых клиентов'],
        traffic_state_data('yt', startdate, enddate, clients),
        traffic_state_data('vk', startdate, enddate, clients),
        traffic_state_data('tk', startdate, enddate, clients),
        traffic_state_data('tv', startdate, enddate, clients),
        traffic_state_data('tg', startdate, enddate, clients),
        traffic_state_data('ref', startdate, enddate, clients)]
    ws = wb.create_sheet("Источники трафика")
    for row in data:
        ws.append(row)


def link_state_data(link):
    category = link.type
    name = link.name
    users = Client.objects.filter(invite_ref=link).count()
    new_buy = link.new_user_buy.count()
    try:
        new_conversion = int(new_buy / users * 100)
    except Exception:
        new_conversion = 0
    new_products_counts = link.new_user_buy.values('name').annotate(count=Count('name')).order_by('-count')
    new_products_counts_str = ''
    for item in new_products_counts:
        new_products_counts_str += f'{item["name"]} {item["count"]}\n'
    old_buy = link.old_user_buy.count()
    old_products_counts = link.old_user_buy.values('name').annotate(count=Count('name')).order_by('-count')
    old_products_counts_str = ''
    for item in old_products_counts:
        old_products_counts_str += f'{item["name"]} {item["count"]}\n'
    return [category, name, users, new_buy, new_conversion, new_products_counts_str, old_buy, old_products_counts_str]


def link_state(wb):
    ws = wb.create_sheet("Статистика ссылок")
    data = [['Категория', 'Название ссылки', 'Количество переходов', 'Количество новых покупок', 'Конверсия',
             'Какие сервисы и в каком колчестве оплачивают новые клиенты', 'Количество старых покупок',
             'Какие сервисы и в каком колчестве оплачивают старые клиенты']]
    for link in ReferralLink.objects.all():
        data.append(link_state_data(link))
    for row in data:
        ws.append(row)


def friend_state_data(startdate, enddate):
    data = []
    for manager in Manager.objects.filter(is_friend=True):
        mailings = manager.individual_mailings.filter(time__range=[startdate, enddate])
        mail_count = mailings.count()
        buy_count = mailings.filter(is_buy=True).count()
        buy_summ = 0
        for mail in mailings:
            buy_summ += mail.buy_summ
        data.append([manager.username, mail_count, buy_count, buy_summ])
    return data


def friend_state(wb, startdate, enddate):
    ws = wb.create_sheet("Статистика МОЗ")
    data = [
               ["Имя МОЗ", "Количество пользователей, которым была сделана рассылка",
                "Количество пользователей, которые купили что-либо в течении трех дней", "Сумма покупок"]
           ] + friend_state_data(startdate, enddate)
    for row in data:
        ws.append(row)


def accounting_state_data(startdate, enddate):
    orders = Order.objects.filter(close_time__range=(startdate, enddate), type__in=['buy', 'not_find_product'],
                                  pay_status='complite')
    data = []
    comission = Comission.objects.get(id=1)
    manager_comission = comission.manager
    cardholder_comission = comission.cardholder
    for order in orders:
        product = order.name
        total_product_price = order.total_product_price
        product_price = order.product_price
        income = total_product_price - product_price
        manager_money = round(
            (float(order.total_product_price) - float(order.product_price)) * (manager_comission / 100), 2)
        ch = Manager.objects.filter(id=order.card_holder_id).first()
        if ch:
            card_holder = ch.username
            cardholder_money = round(
                (float(order.total_product_price) - float(order.product_price)) * (cardholder_comission / 100), 2)
        else:
            card_holder = 'Нет'
            cardholder_money = 0
        true_income = float(income) - manager_money - cardholder_money
        payment_type = order.payment_type
        if order.manager:
            manager = order.manager.username
        else:
            manager = 'Отсутствует'
        data.append([product, total_product_price, product_price, income, manager_money, cardholder_money, true_income, payment_type, card_holder, manager])
    return data


def accounting_state(wb, startdate, enddate):
    ws = wb.create_sheet("Бухгалтерия")

    data = [
        ['Сервисы для оплаты', 'Сделка (сумма)', 'Себестоимость', 'Доход', 'Доля МОП', 'Доля ДК', 'Чистый доход',
         'Способ оплаты', 'Имя ДК', 'Менеджер']
    ] + accounting_state_data(startdate, enddate)

    for row in data:
        ws.append(row)
